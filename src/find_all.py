import requests
from bs4 import BeautifulSoup
from googlesearch import search
import re
import openpyxl
import time

# Function to fetch content from a URL
def fetch_content(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.text
        else:
            print(f"Error {response.status_code} while fetching {url}")
            return ""
    except Exception as e:
        print(f"Error fetching {url}: {e}")
        return ""

# Function to extract email addresses from content
def extract_emails(content):
    return re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", content)

# Function to perform web searches for DevOps job openings
def search_job_postings(queries, num_results=10):
    print("Searching for...")
    urls = []
    for query in queries:
        for idx, result in enumerate(search(query, num_results=num_results)):
            if result not in urls:  # Avoid duplicates
                urls.append(result)
                print(f"Found: {result}")
            if len(urls) >= num_results:  # Stop when enough URLs are collected
                break
            time.sleep(1)  # Add delay to avoid rate-limiting
    return urls

# Function to save email addresses to Excel
def save_to_excel(emails, filename="emails.xlsx"):
    print("Saving to Excel file...")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Emails"
    
    # Header
    sheet.cell(row=1, column=1, value="Email")
    
    # Writing emails to the Excel file
    for idx, email in enumerate(emails, start=2):
        sheet.cell(row=idx, column=1, value=email)
    
    workbook.save(filename)
    print(f"Saved to {filename}")

# Main script
def main():
    # List of search queries
    queries = [
        '"DevOps job opening" "send your resume"',
        '"We are hiring for DevOps" "contact us"',
        '"DevOps engineer required" "apply now"',
        '"Hiring for DevOps" "email us"',
        '"DevOps job vacancies" "apply here"',
        '"Open DevOps positions" "reach out"'
    ]
    
    # Perform searches and collect URLs
    urls = search_job_postings(queries, num_results=50)  # Broaden search scope
    
    all_emails = set()
    for url in urls:
        print(f"Scraping page: {url}")
        content = fetch_content(url)
        
        # Extracting emails directly from the page content
        emails = extract_emails(content)
        if emails:
            print(f"Emails found: {emails}")
        all_emails.update(emails)
    
    if all_emails:
        save_to_excel(all_emails)
        print("Emails successfully extracted and saved!")
    else:
        print("No emails found.")

if __name__ == "__main__":
    main()
