import requests
from bs4 import BeautifulSoup
from googlesearch import search
import re
import openpyxl
import time
import os
import json
import random

# Function to fetch content from a URL
config_file = rf"../input/config.json"
querry_file = rf"../input/querry.json"


def read_json(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as config:
            configuration = json.load(config)
        return configuration
    except FileNotFoundError:
        print(f"File not found: {config_file}")
        return None

def fetch_content(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.text
        else:
            print(f"Error {response.status_code} while fetching {url}")
            return ""
    except Exception as e:
        print(f"Error fetching {url}: {e}")
        return ""

# Function to extract email addresses from content
def extract_emails(content):
    return re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", content)

# Function to perform web searches for DevOps job openings
def search_job_postings(queries, num_results=10):
    print("Searching for Querry...")
    urls = []
    for query in queries:
        for idx, result in enumerate(search(query, num_results=num_results)):
            if result not in urls:  # Avoid duplicates
                urls.append(result)
                print(f"Found: {result}")
            if len(urls) >= num_results:  # Stop when enough URLs are collected
                break
    return urls

# Function to save email addresses and their sources to Excel
def save_to_excel(email_sources, filename="general_devops_2024.xlsx"):
    print("Saving to Excel file...")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Emails and Sources"
    
    # Headers
    sheet.cell(row=1, column=1, value="Email")
    sheet.cell(row=1, column=2, value="Source URL")
    
    # Writing emails and sources to the Excel file
    for idx, (email, source) in enumerate(email_sources, start=2):
        sheet.cell(row=idx, column=1, value=email)
        sheet.cell(row=idx, column=2, value=source)
    
    workbook.save(filename)
    print(f"Saved to {filename}")

# Main script
def main():
    search_querry = read_json(querry_file)
    # List of search queries
    queries = search_querry["queries"]
    
    # Perform searches and collect URLs
    urls = search_job_postings(queries, num_results=50)  # Broaden search scope
    
    email_sources = set()
    for url in urls:
        print(f"Scraping page: {url}")
        
        content = fetch_content(url)

        delay_seconds = random.uniform(2, 5)
        print(f"Sleeping for {delay_seconds:.2f} seconds...")
        time.sleep(delay_seconds)
    
        # Extracting emails directly from the page content
        emails = extract_emails(content)
        if emails:
            print(f"Emails found: {emails}")
            for email in emails:
                if not email.endswith('.png'):
                    email_sources.add((email, url))  # Add email and its source URL
    
    if email_sources:
        save_to_excel(email_sources)
        print("Emails successfully extracted and saved!")
    else:
        print("No emails found.")

if __name__ == "__main__":
    main()
